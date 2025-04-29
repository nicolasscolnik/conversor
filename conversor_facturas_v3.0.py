#!/usr/bin/env python3
import os
import re
import json
import time
import uuid
import requests
import base64
import pdfplumber
import customtkinter as ctk
import tkinter as tk
from io import BytesIO
from pdf2image import convert_from_path
from tkinter import filedialog, messagebox
from tkinterdnd2 import DND_FILES, TkinterDnD
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from dotenv import load_dotenv

# ---------- CONFIGURACIÓN ----------
load_dotenv()  # 🚀 Cargar variables de entorno
TIPOS_FACTURA = {
    '01': 'A', '06': 'B', '11': 'C', '51': 'M', '19': 'E', '91': 'T'
}
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
FOLDER_BASE_DRIVE = '1GJ7h4RgKIU9mBvf_PrBfs1_tL5XFv2L4'
# -------------------------------------

class FacturaExtractorApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Gestor de Facturas")
        self.master.geometry("900x700")
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.setup_ui()
        self.pdf_files = []
        self.data = []

    def setup_ui(self):
        frame = ctk.CTkFrame(self.master)
        frame.pack(padx=20, pady=20, fill="both", expand=True)

        self.button_select = ctk.CTkButton(frame, text="Seleccionar Carpeta", command=self.select_folder)
        self.button_select.pack(pady=10)

        self.drag_drop_label = tk.Label(frame, text="Arrastre aquí sus archivos PDF", relief="groove", height=10)
        self.drag_drop_label.pack(pady=10, fill="both", expand=True)
        self.drag_drop_label.drop_target_register(DND_FILES)
        self.drag_drop_label.dnd_bind('<<Drop>>', self.drop)

        self.button_process = ctk.CTkButton(frame, text="Procesar Facturas", command=self.process_pdfs)
        self.button_process.pack(pady=10)

        self.progress = ttk.Progressbar(frame, orient="horizontal", mode="determinate", length=400)
        self.progress.pack(pady=10)

        self.preview_text = tk.Text(frame, height=15)
        self.preview_text.pack(pady=10, fill="both", expand=True)

    def select_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.pdf_files.clear()
            for root, dirs, files in os.walk(folder_selected):
                for filename in files:
                    if filename.lower().endswith('.pdf'):
                        full_path = os.path.join(root, filename)
                        self.pdf_files.append(full_path)
            if self.pdf_files:
                messagebox.showinfo("Archivos seleccionados", f"{len(self.pdf_files)} archivos PDF encontrados.")
            else:
                messagebox.showwarning("Advertencia", "No se encontraron archivos PDF en la carpeta.")
        else:
            messagebox.showwarning("Advertencia", "No se seleccionó ninguna carpeta.")

    def drop(self, event):
        raw_data = event.data
        files = []
        current = ""
        inside_brace = False

        for c in raw_data:
            if c == '{':
                inside_brace = True
                current = ""
            elif c == '}':
                inside_brace = False
                files.append(current)
            elif c == ' ' and not inside_brace:
                if current:
                    files.append(current)
                    current = ""
            else:
                current += c

        if current:
            files.append(current)

        cleaned_files = []
        for f in files:
            if f.lower().endswith('.pdf') and os.path.isfile(f):
                cleaned_files.append(f)

        if cleaned_files:
            self.pdf_files.extend(cleaned_files)
            messagebox.showinfo("Archivos arrastrados", f"{len(cleaned_files)} archivos PDF agregados.")
        else:
            messagebox.showwarning("Advertencia", "No se detectaron archivos PDF válidos.")

    def process_pdfs(self):
        if not self.pdf_files:
            messagebox.showwarning("Advertencia", "Primero seleccione o arrastre archivos PDF.")
            return

        self.data.clear()
        self.progress["maximum"] = len(self.pdf_files)
        self.progress["value"] = 0

        for pdf_path in self.pdf_files:
            resultado = self.send_pdf_to_openai(pdf_path)
            if resultado:
                self.data.append(resultado)

            self.progress["value"] += 1
            self.master.update()
            time.sleep(0.1)

        if not self.data:
            messagebox.showerror("Error", "No se pudo obtener respuesta de OpenAI.")
            return

        self.export_to_excel()
        self.mostrar_resumen()

        self.progress["value"] = 0 
        self.pdf_files.clear()

    def send_pdf_to_openai(self, pdf_path):
        headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json"
        }

        try:
            images = convert_from_path(pdf_path, dpi=200, first_page=1, last_page=1)
            if not images:
                raise Exception("No se pudo convertir el PDF a imagen.")

            buffered = BytesIO()
            images[0].save(buffered, format="JPEG")
            img_base64 = base64.b64encode(buffered.getvalue()).decode()
            image_url = f"data:image/jpeg;base64,{img_base64}"

            response = requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers=headers,
                json={
                    "model": "gpt-4o",
                    "messages": [
                        {
                            "role": "system",
                            "content": "Sos un experto lector de facturas en imagen. Extraé:\n"
                                       "- Fecha\n- Tipo de Factura (A, B, C, M, E, T)\n"
                                       "- Número de Factura\n- Nombre del Vendedor\n- Monto Total\n"
                                       "Devolveme solo un JSON."
                        },
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": "Extrae los datos de esta factura."},
                                {"type": "image_url", "image_url": {"url": image_url}}
                            ]
                        }
                    ],
                    "temperature": 0,
                    "max_tokens": 1000
                }
            )

            response.raise_for_status()

            result_text = response.json()['choices'][0]['message']['content']

            result_text = result_text.strip()
            if result_text.startswith("```json"):
                result_text = result_text[7:]
            if result_text.endswith("```"):
                result_text = result_text[:-3]
            result_text = result_text.strip()

            return json.loads(result_text)

        except Exception as e:
            print(f"❌ Error procesando {pdf_path}: {e}")
            return None

    def export_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"

        headers = ["Fecha", "Tipo de Factura", "Número de Factura", "Nombre del Vendedor", "Monto Total"]
        ws.append(headers)

        rojo = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        for entry in self.data:
            fila = []
            for header in headers:
                key = header
                value = entry.get(key, "NO DETECTADO")
                if header == "Monto Total" and value != "NO DETECTADO":
                    value = str(value).replace("$", "").replace(" ", "").replace(".", "").replace(",", ".")
                    try:
                        value = float(value)
                    except Exception:
                        value = "NO DETECTADO"
                fila.append(value)
            ws.append(fila)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                if isinstance(cell.value, str) and "NO DETECTADO" in cell.value:
                    cell.fill = rojo

        for col_num, _ in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 25

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'

        fecha_hoy = datetime.now().strftime("%Y-%m-%d")
        nombre_archivo_base = f"Resultados_Facturas_{fecha_hoy}.xlsx"
        nombre_archivo = nombre_archivo_base

        contador = 1
        while os.path.exists(nombre_archivo):
            nombre_archivo = f"Resultados_Facturas_{fecha_hoy} ({contador}).xlsx"
            contador += 1

        wb.save(nombre_archivo)
        print(f"✅ Archivo guardado exitosamente como: {nombre_archivo}")

    def mostrar_resumen(self):
        total_facturas = len(self.data)
        facturas_con_error = sum(1 for entry in self.data if any(v == "NO DETECTADO" for v in entry.values()))
        monto_total = 0.0

        for entry in self.data:
            monto = entry.get("Monto Total")
            if monto and monto != "NO DETECTADO":
                try:
                    monto_clean = str(monto).replace('.', '').replace(',', '.')
                    monto_total += float(monto_clean)
                except Exception:
                    continue

        mensaje = (
            f"Facturas procesadas: {total_facturas}\n"
            f"Facturas con errores: {facturas_con_error}\n"
            f"Total facturado: ${monto_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        )

        messagebox.showinfo("Resumen de procesamiento", mensaje)

if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = FacturaExtractorApp(root)
    root.mainloop()
